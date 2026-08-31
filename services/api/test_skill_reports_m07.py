from io import BytesIO

from openpyxl import load_workbook

import seed_all
from conftest import TestingSessionLocal
from db.models import AssessmentSession, Attempt, AttemptResponse, AuditLog, ContentItem, Student


def _seed_one_graded_response():
    seed_all.run_seed_all()
    db = TestingSessionLocal()
    try:
        student = db.query(Student).filter(Student.access_code == "STU001").one()
        item = (
            db.query(ContentItem)
            .filter(ContentItem.status == "approved")
            .order_by(ContentItem.id)
            .first()
        )
        assert item is not None
        assert item.steps
        session = AssessmentSession(
            student_id=student.id,
            session_type="pretest",
            status="in_progress",
            assigned_level=1,
        )
        db.add(session)
        db.flush()
        attempt = Attempt(
            session_id=session.id,
            item_id=item.id,
            status="in_progress",
        )
        db.add(attempt)
        db.flush()
        db.add(AttemptResponse(
            attempt_id=attempt.id,
            step_id=item.steps[0].id,
            selected_option_id=None,
            is_correct=True,
            elapsed_seconds=4,
        ))
        db.commit()
        return student.id, item.skill.canonical_skill_id or item.skill.skill_key
    finally:
        db.close()


def test_skill_report_uses_persisted_graded_response_evidence_only(researcher_client):
    student_id, skill_code = _seed_one_graded_response()

    response = researcher_client.get(f"/researcher/reports/students/{student_id}/skills")
    assert response.status_code == 200
    payload = response.json()
    assert payload["is_mastery_score"] is False
    row = next(item for item in payload["skills"] if item["skill_code"] == skill_code)
    assert row["graded_responses"] == 1
    assert row["correct_responses"] == 1
    assert row["incorrect_responses"] == 0
    assert row["observed_accuracy_percent"] == 100.0
    assert row["evidence_scope"] == "persisted_graded_responses_only"
    assert row["is_mastery_score"] is False


def test_cohort_skill_report_explicitly_separates_reporting_from_mastery(researcher_client):
    _seed_one_graded_response()

    response = researcher_client.get("/researcher/reports/skills")
    assert response.status_code == 200
    payload = response.json()
    assert payload["cohort_skills"]
    assert all(row["is_mastery_score"] is False for row in payload["cohort_skills"])
    assert "does not change placement" in payload["methodology"]["academic_effect"]
    assert "No speech-derived" in payload["methodology"]["speech"]


def test_skill_xlsx_export_is_valid_and_audited(researcher_client):
    _seed_one_graded_response()

    response = researcher_client.get("/researcher/reports/exports/skills.xlsx")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["المهارات", "المنهجية"]
    assert workbook["المهارات"].max_row >= 2

    db = TestingSessionLocal()
    try:
        audit = (
            db.query(AuditLog)
            .filter(AuditLog.entity_type == "research_skill_evidence")
            .order_by(AuditLog.id.desc())
            .first()
        )
        assert audit is not None
        assert audit.action == "research_report_export"
        assert '"format": "xlsx"' in (audit.details or "")
    finally:
        db.close()


def test_student_skill_report_requires_existing_student(researcher_client):
    response = researcher_client.get("/researcher/reports/students/99999/skills")
    assert response.status_code == 404
    assert response.json()["detail"] == "الطالب غير موجود"
