"""Descriptive per-skill research evidence for M07.

This module summarizes only persisted graded responses. It deliberately does not
calculate mastery, placement, promotion, or any new academic score. The summary
is reporting evidence only and therefore cannot change the student's journey.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import json

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from db.activity_models import ActivityStepResponse
from db.models import (
    AssessmentSession,
    Attempt,
    AttemptResponse,
    AuditLog,
    ContentItem,
    Skill,
    Student,
    User,
)
from dependencies import get_current_user, get_db

router = APIRouter(prefix="/researcher/reports", tags=["Research Reports"])


def _student_attempt_skill_index(db: Session, student_id: int) -> dict[int, Skill]:
    rows = (
        db.query(Attempt.id, Skill)
        .join(AssessmentSession, Attempt.session_id == AssessmentSession.id)
        .join(ContentItem, Attempt.item_id == ContentItem.id)
        .join(Skill, ContentItem.skill_id == Skill.id)
        .filter(AssessmentSession.student_id == student_id)
        .all()
    )
    return {int(attempt_id): skill for attempt_id, skill in rows}


def build_student_skill_evidence(db: Session, student_id: int) -> list[dict]:
    """Return descriptive graded-response counts grouped by canonical skill.

    Every persisted graded response is an observation. Multiple retry rows remain
    visible as multiple observations; this endpoint intentionally does not apply
    the 50/30/20 mastery weighting or any promotion/adaptation rule.
    """
    index = _student_attempt_skill_index(db, student_id)
    if not index:
        return []

    attempt_ids = list(index)
    aggregates: dict[tuple[int, str, str], dict[str, int | str]] = {}

    def add_observation(attempt_id: int, is_correct: bool) -> None:
        skill = index.get(int(attempt_id))
        if skill is None:
            return
        skill_code = str(skill.canonical_skill_id or skill.skill_key)
        key = (int(skill.level_id), skill_code, str(skill.name))
        row = aggregates.setdefault(
            key,
            {
                "level": int(skill.level_id),
                "skill_code": skill_code,
                "skill_name": str(skill.name),
                "graded_responses": 0,
                "correct_responses": 0,
                "incorrect_responses": 0,
            },
        )
        row["graded_responses"] = int(row["graded_responses"]) + 1
        bucket = "correct_responses" if is_correct else "incorrect_responses"
        row[bucket] = int(row[bucket]) + 1

    choice_rows = (
        db.query(AttemptResponse.attempt_id, AttemptResponse.is_correct)
        .filter(
            AttemptResponse.attempt_id.in_(attempt_ids),
            AttemptResponse.is_correct.isnot(None),
        )
        .all()
    )
    for attempt_id, is_correct in choice_rows:
        add_observation(int(attempt_id), bool(is_correct))

    activity_rows = (
        db.query(ActivityStepResponse.attempt_id, ActivityStepResponse.is_correct)
        .filter(ActivityStepResponse.attempt_id.in_(attempt_ids))
        .all()
    )
    for attempt_id, is_correct in activity_rows:
        add_observation(int(attempt_id), bool(is_correct))

    result: list[dict] = []
    for key in sorted(aggregates):
        row = dict(aggregates[key])
        total = int(row["graded_responses"])
        correct = int(row["correct_responses"])
        row["observed_accuracy_percent"] = round((correct / total) * 100.0, 2) if total else None
        row["evidence_scope"] = "persisted_graded_responses_only"
        row["is_mastery_score"] = False
        result.append(row)
    return result


def build_cohort_skill_evidence(db: Session) -> dict:
    students = db.query(Student).order_by(Student.id).all()
    student_rows = []
    combined: dict[tuple[int, str, str], dict[str, int | str]] = {}

    for student in students:
        skills = build_student_skill_evidence(db, student.id)
        student_rows.append({
            "student_id": student.id,
            "student_name": student.name,
            "skills": skills,
        })
        for skill in skills:
            key = (int(skill["level"]), str(skill["skill_code"]), str(skill["skill_name"]))
            row = combined.setdefault(
                key,
                {
                    "level": key[0],
                    "skill_code": key[1],
                    "skill_name": key[2],
                    "graded_responses": 0,
                    "correct_responses": 0,
                    "incorrect_responses": 0,
                },
            )
            for field in ("graded_responses", "correct_responses", "incorrect_responses"):
                row[field] = int(row[field]) + int(skill[field])

    cohort_rows = []
    for key in sorted(combined):
        row = dict(combined[key])
        total = int(row["graded_responses"])
        correct = int(row["correct_responses"])
        row["observed_accuracy_percent"] = round((correct / total) * 100.0, 2) if total else None
        row["evidence_scope"] = "persisted_graded_responses_only"
        row["is_mastery_score"] = False
        cohort_rows.append(row)

    return {
        "cohort_skills": cohort_rows,
        "students": student_rows,
        "methodology": {
            "source": "Persisted AttemptResponse and ActivityStepResponse rows with explicit grading evidence.",
            "retry_policy": "Retries remain separate observations; no 50/30/20 mastery weighting is applied in research reporting.",
            "academic_effect": "Descriptive reporting only; this endpoint does not change placement, mastery, reinforcement, promotion, or posttest eligibility.",
            "speech": "No speech-derived pronunciation categories are emitted before calibrated speech evidence exists.",
        },
    }


def _skills_xlsx_bytes(payload: dict) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "المهارات"
    sheet.sheet_view.rightToLeft = True
    headers = [
        "المستوى",
        "رمز المهارة",
        "المهارة",
        "الاستجابات المقيمة",
        "صحيح",
        "غير صحيح",
        "الدقة المرصودة %",
    ]
    sheet.append(headers)
    for row in payload["cohort_skills"]:
        sheet.append([
            row["level"],
            row["skill_code"],
            row["skill_name"],
            row["graded_responses"],
            row["correct_responses"],
            row["incorrect_responses"],
            row["observed_accuracy_percent"],
        ])

    fill = PatternFill("solid", fgColor="347FD9")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="right", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [12, 30, 30, 20, 14, 14, 20]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    notes = workbook.create_sheet("المنهجية")
    notes.sheet_view.rightToLeft = True
    notes.append(["البند", "التوضيح"])
    for key, value in payload["methodology"].items():
        notes.append([key, value])
    for cell in notes[1]:
        cell.fill = fill
        cell.font = font
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 100

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _audit_export(db: Session, user: User, export_format: str) -> None:
    db.add(AuditLog(
        actor_role="researcher",
        actor_id=user.id,
        action="research_report_export",
        entity_type="research_skill_evidence",
        entity_id="all",
        details=json.dumps({
            "format": export_format,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "policy": "persisted_graded_responses_descriptive_only",
        }, ensure_ascii=False),
    ))
    db.commit()


@router.get("/skills")
def cohort_skill_evidence(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    del user
    return build_cohort_skill_evidence(db)


@router.get("/students/{student_id}/skills")
def student_skill_evidence(
    student_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    del user
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    return {
        "student_id": student.id,
        "student_name": student.name,
        "skills": build_student_skill_evidence(db, student.id),
        "is_mastery_score": False,
    }


@router.get("/exports/skills.xlsx")
def export_skills_xlsx(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    payload = build_cohort_skill_evidence(db)
    content = _skills_xlsx_bytes(payload)
    _audit_export(db, user, "xlsx")
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="himma-research-skills.xlsx"'},
    )
