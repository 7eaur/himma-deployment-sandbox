"""Research reporting endpoints for M07.

This module reports persisted academic/runtime evidence. It deliberately does not
recalculate placement or adaptation rules, and it does not invent speech-derived
metrics when calibrated speech evidence is unavailable.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import json
import os
from pathlib import Path

import arabic_reshaper
from bidi.algorithm import get_display
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func
from sqlalchemy.orm import Session

from db.models import AssessmentSession, Attempt, AuditLog, Student, User
from db.reinforcement_models import ReinforcementCycle
from dependencies import get_current_user, get_db

router = APIRouter(prefix="/researcher/reports", tags=["Research Reports"])


REPORT_FONT_NAME = "HimmaArabicReport"


def _number(value) -> float | None:
    if value is None:
        return None
    return float(value)


def _improvement(pre_score: float | None, post_score: float | None) -> tuple[float | None, float | None]:
    """Return percentage-point and relative improvements without inventing zero baselines."""
    if pre_score is None or post_score is None:
        return None, None
    absolute = round(post_score - pre_score, 4)
    relative = None if pre_score == 0 else round((absolute / pre_score) * 100.0, 4)
    return absolute, relative


def _sessions_by_type(db: Session, student_id: int) -> dict[str, AssessmentSession | None]:
    rows = db.query(AssessmentSession).filter(
        AssessmentSession.student_id == student_id,
        AssessmentSession.session_type.in_(["pretest", "posttest"]),
    ).order_by(AssessmentSession.id).all()
    result: dict[str, AssessmentSession | None] = {"pretest": None, "posttest": None}
    for row in rows:
        result[row.session_type] = row
    return result


def build_student_research_report(db: Session, student: Student) -> dict:
    sessions = _sessions_by_type(db, student.id)
    pre = sessions["pretest"]
    post = sessions["posttest"]

    pre_score = _number(pre.final_score) if pre and pre.status == "completed" else None
    post_score = _number(post.final_score) if post and post.status == "completed" else None
    absolute_improvement, relative_improvement = _improvement(pre_score, post_score)

    core_sessions = db.query(AssessmentSession).filter(
        AssessmentSession.student_id == student.id,
        AssessmentSession.session_type == "core",
    ).order_by(AssessmentSession.id).all()
    all_session_ids = [session.id for session in [pre, post, *core_sessions] if session is not None]
    attempt_count = 0
    completed_attempt_count = 0
    if all_session_ids:
        attempt_count = db.query(func.count(Attempt.id)).filter(Attempt.session_id.in_(all_session_ids)).scalar() or 0
        completed_attempt_count = db.query(func.count(Attempt.id)).filter(
            Attempt.session_id.in_(all_session_ids),
            Attempt.status == "completed",
        ).scalar() or 0

    cycles = db.query(ReinforcementCycle).filter(
        ReinforcementCycle.student_id == student.id,
    ).order_by(ReinforcementCycle.id).all()
    cycle_counts = {
        "total": len(cycles),
        "verified": sum(1 for cycle in cycles if cycle.status == "verified"),
        "escalated": sum(1 for cycle in cycles if cycle.status == "escalated"),
        "active": sum(1 for cycle in cycles if cycle.status not in {"verified", "escalated"}),
    }

    completed_core_levels = [
        int(session.assigned_level)
        for session in core_sessions
        if session.status == "completed" and session.assigned_level is not None
    ]

    return {
        "student_id": student.id,
        "student_name": student.name,
        "status": "active" if student.is_active else "inactive",
        "starting_level": int(pre.assigned_level) if pre and pre.status == "completed" and pre.assigned_level else None,
        "current_level": student.current_level,
        "final_level": int(post.assigned_level) if post and post.status == "completed" and post.assigned_level else None,
        "completed_core_levels": completed_core_levels,
        "pretest": {
            "status": pre.status if pre else "not_started",
            "score": pre_score,
            "elapsed_seconds": pre.elapsed_seconds if pre else 0,
            "completed_at": pre.completed_at if pre else None,
        },
        "posttest": {
            "status": post.status if post else "not_started",
            "score": post_score,
            "elapsed_seconds": post.elapsed_seconds if post else 0,
            "completed_at": post.completed_at if post else None,
        },
        "improvement": {
            "absolute_percentage_points": absolute_improvement,
            "relative_percent": relative_improvement,
            "relative_percent_defined": relative_improvement is not None,
        },
        "engagement": {
            "assessment_seconds": (pre.elapsed_seconds if pre else 0) + (post.elapsed_seconds if post else 0),
            "learning_seconds": sum(session.elapsed_seconds for session in core_sessions),
            "attempts": int(attempt_count),
            "completed_attempts": int(completed_attempt_count),
        },
        "reinforcement": cycle_counts,
        "speech_evidence": {
            "calibrated": False,
            "error_categories": None,
            "note": "لا تُعرض أخطاء نطق آلية قبل وجود دليل صوتي مُعاير ومعتمد.",
        },
    }


def build_cohort_research_report(db: Session) -> dict:
    students = db.query(Student).order_by(Student.id).all()
    reports = [build_student_research_report(db, student) for student in students]

    paired = [
        report for report in reports
        if report["pretest"]["score"] is not None and report["posttest"]["score"] is not None
    ]
    pre_scores = [report["pretest"]["score"] for report in reports if report["pretest"]["score"] is not None]
    post_scores = [report["posttest"]["score"] for report in reports if report["posttest"]["score"] is not None]
    improvements = [report["improvement"]["absolute_percentage_points"] for report in paired]

    def avg(values: list[float]) -> float | None:
        return round(sum(values) / len(values), 4) if values else None

    return {
        "cohort": {
            "students": len(reports),
            "active_students": sum(1 for report in reports if report["status"] == "active"),
            "completed_pretests": len(pre_scores),
            "completed_posttests": len(post_scores),
            "paired_pre_post": len(paired),
            "average_pretest_score": avg(pre_scores),
            "average_posttest_score": avg(post_scores),
            "average_absolute_improvement_points": avg(improvements),
            "reinforcement_cycles": sum(report["reinforcement"]["total"] for report in reports),
            "verified_reinforcement_cycles": sum(report["reinforcement"]["verified"] for report in reports),
            "escalated_reinforcement_cycles": sum(report["reinforcement"]["escalated"] for report in reports),
        },
        "students": reports,
        "reporting_notes": {
            "score_source": "Persisted completed assessment session scores; M07 does not recalculate placement.",
            "relative_improvement": "((post - pre) / pre) × 100; null when pretest score is 0 or either test is incomplete.",
            "speech_metrics": "Unavailable until calibrated speech evidence is accepted.",
        },
    }


def _audit_export(db: Session, user: User, *, entity_type: str, entity_id: str, export_format: str) -> None:
    db.add(AuditLog(
        actor_role="researcher",
        actor_id=user.id,
        action="research_report_export",
        entity_type=entity_type,
        entity_id=entity_id,
        details=json.dumps({
            "format": export_format,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "policy": "persisted_evidence_only",
        }, ensure_ascii=False),
    ))
    db.commit()


def _score_text(value: float | None) -> str:
    return "—" if value is None else f"{round(value, 1)}%"


def _xlsx_bytes(report: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ملخص"
    ws.sheet_view.rightToLeft = True
    header_fill = PatternFill("solid", fgColor="347FD9")
    header_font = Font(color="FFFFFF", bold=True)

    cohort = report["cohort"]
    summary_rows = [
        ("المؤشر", "القيمة"),
        ("إجمالي الطلاب", cohort["students"]),
        ("الطلاب النشطون", cohort["active_students"]),
        ("الاختبارات القبلية المكتملة", cohort["completed_pretests"]),
        ("الاختبارات البعدية المكتملة", cohort["completed_posttests"]),
        ("مقارنات قبلي/بعدي مكتملة", cohort["paired_pre_post"]),
        ("متوسط القبلي", cohort["average_pretest_score"]),
        ("متوسط البعدي", cohort["average_posttest_score"]),
        ("متوسط التحسن بالنقاط", cohort["average_absolute_improvement_points"]),
        ("دورات التقوية", cohort["reinforcement_cycles"]),
        ("دورات التقوية المتحققة", cohort["verified_reinforcement_cycles"]),
        ("دورات التقوية المصعّدة", cohort["escalated_reinforcement_cycles"]),
    ]
    for row in summary_rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24

    students_ws = wb.create_sheet("الطلاب")
    students_ws.sheet_view.rightToLeft = True
    headers = [
        "الطالب", "الحالة", "مستوى البداية", "المستوى الحالي", "المستوى النهائي",
        "القبلي", "البعدي", "التحسن بالنقاط", "التحسن النسبي %", "زمن الاختبارات ث",
        "زمن التعلم ث", "المحاولات", "المحاولات المكتملة", "التقويات", "المتحقق منها", "المصعّد",
    ]
    students_ws.append(headers)
    for student in report["students"]:
        students_ws.append([
            student["student_name"],
            "نشط" if student["status"] == "active" else "غير نشط",
            student["starting_level"],
            student["current_level"],
            student["final_level"],
            student["pretest"]["score"],
            student["posttest"]["score"],
            student["improvement"]["absolute_percentage_points"],
            student["improvement"]["relative_percent"],
            student["engagement"]["assessment_seconds"],
            student["engagement"]["learning_seconds"],
            student["engagement"]["attempts"],
            student["engagement"]["completed_attempts"],
            student["reinforcement"]["total"],
            student["reinforcement"]["verified"],
            student["reinforcement"]["escalated"],
        ])
    for cell in students_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right", wrap_text=True)
    students_ws.freeze_panes = "A2"
    students_ws.auto_filter.ref = students_ws.dimensions
    for column_index in range(1, len(headers) + 1):
        students_ws.column_dimensions[get_column_letter(column_index)].width = 18 if column_index > 1 else 28

    reinforcement_ws = wb.create_sheet("التقوية")
    reinforcement_ws.sheet_view.rightToLeft = True
    reinforcement_ws.append(["الطالب", "إجمالي الدورات", "متحقق", "مصعّد", "نشط"])
    for student in report["students"]:
        reinforcement_ws.append([
            student["student_name"],
            student["reinforcement"]["total"],
            student["reinforcement"]["verified"],
            student["reinforcement"]["escalated"],
            student["reinforcement"]["active"],
        ])
    for cell in reinforcement_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right")
    reinforcement_ws.column_dimensions["A"].width = 28
    for column in "BCDE":
        reinforcement_ws.column_dimensions[column].width = 18

    notes_ws = wb.create_sheet("ملاحظات منهجية")
    notes_ws.sheet_view.rightToLeft = True
    notes_ws.append(["البند", "الملاحظة"])
    notes_ws.append(["مصدر الدرجات", report["reporting_notes"]["score_source"]])
    notes_ws.append(["التحسن النسبي", report["reporting_notes"]["relative_improvement"]])
    notes_ws.append(["الصوت", report["reporting_notes"]["speech_metrics"]])
    for cell in notes_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    notes_ws.column_dimensions["A"].width = 24
    notes_ws.column_dimensions["B"].width = 90

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _find_report_font() -> Path | None:
    configured = os.getenv("HIMMA_REPORT_FONT_PATH")
    candidates = [
        configured,
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansArabic-Regular.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).is_file():
            return Path(candidate)
    return None


def _rtl(value: object) -> str:
    text = str(value if value is not None else "—")
    return get_display(arabic_reshaper.reshape(text))


def _ensure_pdf_font() -> str:
    if REPORT_FONT_NAME in pdfmetrics.getRegisteredFontNames():
        return REPORT_FONT_NAME
    path = _find_report_font()
    if path is None:
        raise HTTPException(
            status_code=503,
            detail="خط التقارير العربية غير متوفر على الخادم. اضبط HIMMA_REPORT_FONT_PATH قبل تفعيل تصدير PDF.",
        )
    pdfmetrics.registerFont(TTFont(REPORT_FONT_NAME, str(path)))
    return REPORT_FONT_NAME


def _pdf_styles():
    font = _ensure_pdf_font()
    base = getSampleStyleSheet()
    title = ParagraphStyle(
        "HimmaTitle", parent=base["Title"], fontName=font, fontSize=17,
        leading=24, alignment=TA_RIGHT, textColor=colors.HexColor("#20364D"),
    )
    body = ParagraphStyle(
        "HimmaBody", parent=base["BodyText"], fontName=font, fontSize=10,
        leading=16, alignment=TA_RIGHT,
    )
    small = ParagraphStyle(
        "HimmaSmall", parent=body, fontSize=8.5, leading=13,
    )
    return font, title, body, small


def _student_pdf_bytes(student: dict) -> bytes:
    font, title_style, body_style, _ = _pdf_styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title=f"Himma student report {student['student_id']}",
    )
    story = [
        Paragraph(_rtl("تقرير الطالب - منصة هِمّة"), title_style),
        Spacer(1, 8 * mm),
        Paragraph(_rtl(f"الطالب: {student['student_name']}"), body_style),
        Paragraph(_rtl(f"مستوى البداية: {student['starting_level'] or '—'} | المستوى الحالي: {student['current_level']} | المستوى النهائي: {student['final_level'] or '—'}"), body_style),
        Spacer(1, 5 * mm),
    ]
    rows = [
        [_rtl("المؤشر"), _rtl("القيمة")],
        [_rtl("الاختبار القبلي"), _rtl(_score_text(student["pretest"]["score"]))],
        [_rtl("الاختبار البعدي"), _rtl(_score_text(student["posttest"]["score"]))],
        [_rtl("التحسن بالنقاط"), _rtl(student["improvement"]["absolute_percentage_points"])],
        [_rtl("التحسن النسبي"), _rtl(student["improvement"]["relative_percent"])],
        [_rtl("زمن الاختبارات (ث)"), _rtl(student["engagement"]["assessment_seconds"])],
        [_rtl("زمن التعلم (ث)"), _rtl(student["engagement"]["learning_seconds"])],
        [_rtl("المحاولات"), _rtl(student["engagement"]["attempts"])],
        [_rtl("دورات التقوية"), _rtl(student["reinforcement"]["total"])],
        [_rtl("التقويات المتحققة"), _rtl(student["reinforcement"]["verified"])],
    ]
    table = Table(rows, colWidths=[95 * mm, 75 * mm], hAlign="RIGHT")
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#347FD9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DCE8F2")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(table)
    story.append(Spacer(1, 7 * mm))
    story.append(Paragraph(_rtl("ملاحظة: لا تُعرض مؤشرات أخطاء النطق الآلية قبل اعتماد دليل صوتي مُعاير."), body_style))
    doc.build(story)
    return buffer.getvalue()


def _cohort_pdf_bytes(report: dict) -> bytes:
    font, title_style, body_style, small_style = _pdf_styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm, title="Himma cohort research report",
    )
    cohort = report["cohort"]
    story = [
        Paragraph(_rtl("التقرير الإجمالي - منصة هِمّة"), title_style),
        Paragraph(_rtl(
            f"الطلاب: {cohort['students']} | قبلي مكتمل: {cohort['completed_pretests']} | "
            f"بعدي مكتمل: {cohort['completed_posttests']} | مقارنات مكتملة: {cohort['paired_pre_post']}"
        ), body_style),
        Spacer(1, 5 * mm),
    ]
    rows = [[
        _rtl("الطالب"), _rtl("البداية"), _rtl("الحالي"), _rtl("النهائي"),
        _rtl("القبلي"), _rtl("البعدي"), _rtl("التحسن"), _rtl("المحاولات"), _rtl("التقوية"),
    ]]
    for student in report["students"]:
        rows.append([
            _rtl(student["student_name"]),
            _rtl(student["starting_level"]),
            _rtl(student["current_level"]),
            _rtl(student["final_level"]),
            _rtl(_score_text(student["pretest"]["score"])),
            _rtl(_score_text(student["posttest"]["score"])),
            _rtl(student["improvement"]["absolute_percentage_points"]),
            _rtl(student["engagement"]["attempts"]),
            _rtl(f"{student['reinforcement']['verified']}/{student['reinforcement']['total']}"),
        ])
    table = Table(rows, repeatRows=1, colWidths=[50 * mm, 20 * mm, 20 * mm, 20 * mm, 25 * mm, 25 * mm, 25 * mm, 22 * mm, 25 * mm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#347FD9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#DCE8F2")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(_rtl("المؤشرات مبنية على القيم المحفوظة فقط، ولا يعاد احتساب التصنيف داخل التقرير."), small_style))
    doc.build(story)
    return buffer.getvalue()


def _stream_bytes(content: bytes, *, media_type: str, filename: str) -> StreamingResponse:
    return StreamingResponse(
        BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/summary")
def research_report_summary(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    del user
    return build_cohort_research_report(db)


@router.get("/students/{student_id}")
def research_report_student(
    student_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    del user
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    return build_student_research_report(db, student)


@router.get("/exports/cohort.xlsx")
def export_cohort_xlsx(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    report = build_cohort_research_report(db)
    content = _xlsx_bytes(report)
    _audit_export(db, user, entity_type="research_cohort", entity_id="all", export_format="xlsx")
    return _stream_bytes(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="himma-research-cohort.xlsx",
    )


@router.get("/exports/cohort.pdf")
def export_cohort_pdf(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    report = build_cohort_research_report(db)
    content = _cohort_pdf_bytes(report)
    _audit_export(db, user, entity_type="research_cohort", entity_id="all", export_format="pdf")
    return _stream_bytes(content, media_type="application/pdf", filename="himma-research-cohort.pdf")


@router.get("/students/{student_id}/export.pdf")
def export_student_pdf(
    student_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    report = build_student_research_report(db, student)
    content = _student_pdf_bytes(report)
    _audit_export(db, user, entity_type="student", entity_id=str(student_id), export_format="pdf")
    return _stream_bytes(
        content,
        media_type="application/pdf",
        filename=f"himma-student-{student_id}.pdf",
    )
