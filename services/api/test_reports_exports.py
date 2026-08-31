"""M07 export contracts: usable files, persisted evidence only, and Arabic PDF guard."""

from io import BytesIO

from openpyxl import load_workbook

from reports import _cohort_pdf_bytes, _student_pdf_bytes, _xlsx_bytes


def _report():
    student = {
        "student_id": 7,
        "student_name": "طالب تجريبي",
        "status": "active",
        "starting_level": 1,
        "current_level": 2,
        "final_level": 2,
        "completed_core_levels": [1],
        "pretest": {"status": "completed", "score": 40.0, "elapsed_seconds": 120, "completed_at": None},
        "posttest": {"status": "completed", "score": 70.0, "elapsed_seconds": 100, "completed_at": None},
        "improvement": {"absolute_percentage_points": 30.0, "relative_percent": 75.0, "relative_percent_defined": True},
        "engagement": {"assessment_seconds": 220, "learning_seconds": 600, "attempts": 12, "completed_attempts": 10},
        "reinforcement": {"total": 2, "verified": 1, "escalated": 0, "active": 1},
        "speech_evidence": {"calibrated": False, "error_categories": None, "note": "غير معاير"},
    }
    return {
        "cohort": {
            "students": 1,
            "active_students": 1,
            "completed_pretests": 1,
            "completed_posttests": 1,
            "paired_pre_post": 1,
            "average_pretest_score": 40.0,
            "average_posttest_score": 70.0,
            "average_absolute_improvement_points": 30.0,
            "reinforcement_cycles": 2,
            "verified_reinforcement_cycles": 1,
            "escalated_reinforcement_cycles": 0,
        },
        "students": [student],
        "reporting_notes": {
            "score_source": "persisted",
            "relative_improvement": "formula",
            "speech_metrics": "unavailable",
        },
    }


def test_xlsx_export_is_valid_multi_sheet_workbook():
    content = _xlsx_bytes(_report())
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    assert workbook.sheetnames == ["ملخص", "الطلاب", "التقوية", "ملاحظات منهجية"]
    assert workbook["الطلاب"]["A2"].value == "طالب تجريبي"
    assert workbook["الطلاب"]["H2"].value == 30.0
    assert workbook["التقوية"]["C2"].value == 1


def test_student_pdf_export_produces_pdf_bytes():
    content = _student_pdf_bytes(_report()["students"][0])
    assert content.startswith(b"%PDF-")
    assert len(content) > 1000


def test_cohort_pdf_export_produces_pdf_bytes():
    content = _cohort_pdf_bytes(_report())
    assert content.startswith(b"%PDF-")
    assert len(content) > 1000
